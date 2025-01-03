import json
import os
import re
import numpy as np
import xlsxwriter
import openpyxl
import ast
import json
from standardize_qn import standardize_vietnamese


# region Helper functions
def load_sino_nom_similar_dic(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    sino_similar_dict = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
        input_char = row[0]
        similar_chars_str = row[1]
        similar_chars = ast.literal_eval(similar_chars_str)  # Convert string to list
        ordered_chars = [input_char] + similar_chars  # Maintain original order
        sino_similar_dict[input_char] = ordered_chars
    return sino_similar_dict


def load_quoc_ngu_sino_nom_dic(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    qn_sino_dict = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        qn_word = row[0]
        sino_char = row[1]

        if qn_word in qn_sino_dict:
            qn_sino_dict[qn_word].append(sino_char)  # Add to list if it exists
        else:
            qn_sino_dict[qn_word] = [sino_char]  # Create new list if not
    return qn_sino_dict


def load_json_files_from_folder(folder_path):
    json_data_list = []

    # Lấy danh sách tệp và sắp xếp theo chiều dài trước, sau đó theo giá trị số
    filenames = sorted(
        [f for f in os.listdir(folder_path) if f.endswith(".json")],
        key=lambda x: (len(x), x),
    )

    print(filenames)

    for filename in filenames:
        file_path = os.path.join(folder_path, filename)
        try:
            with open(file_path, "r", encoding="utf-8") as file:
                json_data = json.load(file)
                json_data_list.append(json_data)
        except Exception as e:
            print(f"Cannot read file {filename}: {e}")

    return json_data_list


# endregion


# region Alignment functions
def create_ocr_nom_string(paragraph):
    results = []
    for line in paragraph:  # Duyệt từng dòng trong 'ocr_text'
        results.extend(list(line))  # Tách từng ký tự và thêm vào danh sách
    return results


def get_cost(ocr_char, qn_word, sino_similar_dict, qn_sino_dict):
    qn_word = re.sub(r"[^\w\s]", "", qn_word).lower()
    qn_word = standardize_vietnamese(qn_word)

    S1 = sino_similar_dict.get(ocr_char, {ocr_char})
    S2 = qn_sino_dict.get(qn_word, set())
    intersection = [char for char in S1 if char in S2]
    if len(intersection) == 0:
        return 2
    else:
        return 0


def Levenshtein_Align_Line(ocr_sentence, qn_sentence, nom_sino_dict, qn_sino_dict):
    m, n = len(ocr_sentence), len(qn_sentence)
    dp = np.zeros((m + 1, n + 1))

    dp[:, 0] = np.arange(m + 1)
    dp[0, :] = np.arange(n + 1)

    # Tính toán chi phí chỉnh sửa
    for i in range(1, m + 1):
        for j in range(1, n + 1):
            cost = get_cost(
                ocr_sentence[i - 1], qn_sentence[j - 1], nom_sino_dict, qn_sino_dict
            )
            dp[i, j] = min(dp[i - 1, j] + 1, dp[i, j - 1] + 1, dp[i - 1, j - 1] + cost)

    aligned_result_ocr = []
    aligned_result_qn = []

    i, j = m, n
    while i > 0 or j > 0:
        if i > 0 and j > 0:
            cost = get_cost(
                ocr_sentence[i - 1], qn_sentence[j - 1], nom_sino_dict, qn_sino_dict
            )
            if dp[i, j] == dp[i - 1, j - 1] + cost:
                status = "red" if cost >= 1 else "black"
                aligned_result_ocr.append((ocr_sentence[i - 1], status))
                aligned_result_qn.append((qn_sentence[j - 1], status))
                i, j = i - 1, j - 1
                continue

        if i > 0 and (j == 0 or dp[i, j] == dp[i - 1, j] + 1):
            aligned_result_ocr.append((ocr_sentence[i - 1], "red"))
            aligned_result_qn.append(("_", "red"))
            i -= 1
        elif j > 0:  # Chèn trong OCR
            aligned_result_ocr.append(("_", "red"))
            aligned_result_qn.append((qn_sentence[j - 1], "red"))
            j -= 1

    return aligned_result_ocr[::-1], aligned_result_qn[::-1]  # Đảo ngược để đúng thứ tự


# endregion


# region Helper function to load a single JSON file
def load_single_json_file(file_path):
    with open(file_path, "r", encoding="utf-8") as f:
        return json.load(f)


def load_single_nom_page(file_path):
    nom_sentences = []
    nom_paragraph = load_single_json_file(file_path)

    if "result" in nom_paragraph:
        for item in nom_paragraph["result"]:
            if "transcription" in item:
                nom_sentences.append(item["transcription"])
    return nom_sentences


def load_single_qngu_page(file_path):
    qn_sentences = []
    qn_paragraph = load_single_json_file(file_path)

    if "ocr_text" in qn_paragraph:
        for line in qn_paragraph["ocr_text"]:
            qn_sentences.extend(line.split())
    return qn_sentences


def load_box_nom_page(file_path, page):
    boxes = []
    data = load_single_json_file(file_path)
    box_number = 1

    if "result" in data:
        for item in data["result"]:
            if "points" in item:
                boxes.append(
                    {"page": page, "point": item["points"], "box_number": box_number}
                )
                box_number += 1
    return boxes


# endregion


# function to cut the qngu redundant => thiếu nôm
def cutRedundantQN(index_last_box, aligned_nom, aligned_qn):
    index = len(aligned_nom) - 1
    last_box_nom = [char for char, status in aligned_nom[index_last_box:]]
    last_box_qn = [char for char, status in aligned_qn[index_last_box:]]

    # replace multiple '_' (from 3 or more) by ''
    last_box_nom = "".join(last_box_nom)
    last_box_nom = re.sub(r"_{3,}", "", last_box_nom)
    print(last_box_nom)
    number_of_nom_char = len(re.findall(r"[^_]", last_box_nom))
    last_box_nom = list(last_box_nom)
    # count number of non '_' character
    last_box_qn = [char for char in last_box_qn if char != "_"]

    redundant_qn = [char for char in last_box_qn[number_of_nom_char:]]

    nom_chars = [char for char in last_box_nom if char != "_"]
    qn_chars = [char for char in last_box_qn[:number_of_nom_char]]

    new_last_box_nom, new_last_box_qn = Levenshtein_Align_Line(
        nom_chars, qn_chars, sino_similar_dict, qn_sino_dict
    )

    return (
        aligned_nom[:index_last_box] + new_last_box_nom,
        aligned_qn[:index_last_box] + new_last_box_qn,
        redundant_qn,
    )


def cutRedundantNom(aligned_nom, aligned_qn, boxes, nom_sentences):
    current_char_index = len(aligned_nom) - 1
    current_box_length = 0
    accumulated_qn = 0
    accumulated_nom = 0
    THRESHOLD = 0.6
    BOXES_THRESHOLD = 3
    NUMBER_OF_WORDS_IN_BOX_THRESHOLD = 3
    count_box_over_threshold = 0
    print("last_box: ", nom_sentences[-1])
    for index in range(len(nom_sentences) - 1, -1, -1):
        print(index)
        end_of_box = current_char_index
        number_of_empty_qn = 0
        current_box_length = len(nom_sentences[index])
        while current_box_length > 0:
            if aligned_nom[current_char_index][0] != "_":
                current_box_length -= 1
                accumulated_nom += 1

            if aligned_qn[current_char_index][0] == "_":
                number_of_empty_qn += 1
            else:
                accumulated_qn += 1

            current_char_index -= 1

        print("troll ", end_of_box - current_char_index + 1)
        percentage = number_of_empty_qn / (end_of_box - current_char_index + 1)
        if percentage < THRESHOLD:
            print('Tien', end_of_box - current_char_index + 1)
            if end_of_box - current_char_index + 1 < NUMBER_OF_WORDS_IN_BOX_THRESHOLD:
                continue
            else:
                break
    
    while aligned_nom[current_char_index][0] == "_":
        current_char_index -= 1
        if aligned_qn[current_char_index][0] != "_":
            accumulated_qn += 1

    print("diff", accumulated_qn, accumulated_nom)
    return aligned_nom[:current_char_index + 1], aligned_qn[:current_char_index + 1], aligned_nom[current_char_index + 1:], aligned_qn[current_char_index + 1:], accumulated_nom - accumulated_qn


# check the last box is redundant:
def isLackNom(aligned_nom, aligned_qn, nom_sentences):
    current_char_index = len(aligned_nom) - 1
    current_box_length = 0
    accumulated_qn = 0
    accumulated_nom = 0
    THRESHOLD = 0.65
    NUMBER_OF_WORDS_IN_BOX_THRESHOLD = 3
    box_count = 0
    print("last_box: ", nom_sentences[-1])
    for index in range(len(nom_sentences) - 1, -1, -1):
        end_of_box = current_char_index
        number_of_empty_qn = 0
        number_of_empty_nom = 0
        current_box_length = len(nom_sentences[index])
        while current_box_length > 0:
            if aligned_nom[current_char_index][0] != "_":
                current_box_length -= 1
                accumulated_nom += 1
            else: 
                number_of_empty_nom += 1

            if aligned_qn[current_char_index][0] == "_":
                number_of_empty_qn += 1
            else:
                accumulated_qn += 1

            current_char_index -= 1

        percentage_of_empty_qn = number_of_empty_qn / (end_of_box - current_char_index + 1)
        percentage_of_empty_nom = number_of_empty_nom / (end_of_box - current_char_index + 1)
        if (percentage_of_empty_qn < THRESHOLD) and (percentage_of_empty_nom < THRESHOLD):
            if end_of_box - current_char_index + 1 < NUMBER_OF_WORDS_IN_BOX_THRESHOLD:
                continue
            else:
                break
        box_count += 1
    
    while aligned_nom[current_char_index][0] == "_":
        current_char_index -= 1
        if aligned_qn[current_char_index][0] != "_":
            accumulated_qn += 1

    print("diff", accumulated_qn, accumulated_nom, end_of_box, current_char_index)
    print("box count: ", box_count)
    return accumulated_qn - accumulated_nom, current_char_index

# region Write to Excel function


def write_to_excel(
    input_nom_folder,
    input_qngu_folder,
    book_name,
    book_page_ranges,
    sino_similar_dict,
    qn_sino_dict,
    output_file,
):
    workbook = xlsxwriter.Workbook(output_file)
    worksheet = workbook.add_worksheet()

    # Define formatting styles
    black_text_format = workbook.add_format(
        {"font_name": "Nom Na Tong", "font_size": 12, "color": "black"}
    )
    header_format = workbook.add_format(
        {"font_name": "Nom Na Tong", "font_size": 12, "bold": True, "align": "center"}
    )
    red_text_format = workbook.add_format(
        {"font_name": "Nom Na Tong", "color": "#FF0000"}
    )

    # Set column widths
    worksheet.set_column("A:A", 50)
    worksheet.set_column("B:B", 40)
    worksheet.set_column("C:C", 40)
    worksheet.set_column("D:D", 80)
    worksheet.set_column("E:E", 120)

    headers = ["Image_name", "ID", "Image Box", "SinoNom OCR", "Chữ Quốc Ngữ"]
    for col, header in enumerate(headers):
        worksheet.write(0, col, header, header_format)

    row_index_global = 1
    total_word = 0
    red_word = 0
    empty_word = 0

    def write_line(row_idx, nom_pairs, qn_pairs, length_ocr_line):
        if length_ocr_line == 1:
            worksheet.write(row_idx, 3, nom_pairs[1], nom_pairs[0])
            worksheet.write(row_idx, 4, qn_pairs[1], qn_pairs[0])
        else:
            worksheet.write_rich_string(row_idx, 3, *nom_pairs)
            worksheet.write_rich_string(row_idx, 4, *qn_pairs)

    redundant_qn = []
    loaded_qn = set()
    loaded_nom = set()
    for i, (pages_x, pages_y) in enumerate(book_page_ranges):
        print(f"Processing page {pages_x} and {pages_y}...")
        boxes = []
        nom_sentences = []
        
        qn_chars = list(redundant_qn)
        for page_x in pages_x:
            if page_x not in loaded_qn:
                qn_chars.extend(
                    load_single_qngu_page(f"{input_qngu_folder}/page_{page_x}.json")
                )
                loaded_qn.add(page_x)

        for page_y in pages_y:
            if page_y not in loaded_nom:
                print(page_y)
                tmp = load_box_nom_page(f"{input_nom_folder}/page_{page_y}.json", page_y)
                boxes.extend(tmp)
                nom_sentences.extend(
                    load_single_nom_page(f"{input_nom_folder}/page_{page_y}.json")
                )
                loaded_nom.add(page_y)

        nom_chars = create_ocr_nom_string(nom_sentences)

        aligned_nom, aligned_qn = Levenshtein_Align_Line(
            nom_chars, qn_chars, sino_similar_dict, qn_sino_dict
        )

        print(pages_x, pages_y)
        index_reverse = len(nom_sentences) - 1

        is_lack_nom, index_last_box = isLackNom(aligned_nom, aligned_qn, nom_sentences)
    
        if is_lack_nom > 0:
            print("Vai nho")
            aligned_nom, aligned_qn, redundant_qn = cutRedundantQN(
                index_last_box, aligned_nom, aligned_qn
            )
        elif is_lack_nom < 0:
            print("Vai lon")
            aligned_nom, aligned_qn, re_align_nom, re_align_qn, diff = cutRedundantNom(
                aligned_nom, aligned_qn, boxes, nom_sentences
            )
            if (i < len(book_page_ranges) - 1):
                next_page_qn = load_single_qngu_page(f"{input_qngu_folder}/page_{book_page_ranges[i + 1][0][0]}.json")
                loaded_qn.add(book_page_ranges[i + 1][0][0])
                redundant_qn = next_page_qn[diff:]

                re_align_nom_chars = [char for char, status in re_align_nom if char != "_"]
                re_align_qn_chars = [char for char, status in re_align_qn if char != "_"] + next_page_qn[:diff]
                re_align_nom, re_align_qn = Levenshtein_Align_Line(
                    re_align_nom_chars, re_align_qn_chars, sino_similar_dict, qn_sino_dict
                )
                aligned_nom.extend(re_align_nom)
                aligned_qn.extend(re_align_qn)
        else:
            print("Khong thieu")

        nom_pairs = []
        qn_pairs = []
        index_ocr_line = 0
        index_ocr_word = 0
        count_char = 0
        length_char = len(aligned_nom)

        while count_char < length_char:
            nom_char, nom_status = aligned_nom[count_char]
            qn_char, qn_status = aligned_qn[count_char]
            length_ocr_line = len(nom_sentences[index_ocr_line])

            nom_pairs.append(
                red_text_format if nom_status == "red" else black_text_format
            )
            nom_pairs.append(nom_char)
            qn_pairs.append(
                red_text_format if qn_status == "red" else black_text_format
            )
            qn_pairs.append(qn_char + " ")

            if nom_char == "_" or qn_char == "_":
                empty_word += 1
            elif nom_status == "red" and qn_status == "red":
                red_word += 1
            total_word += 1

            count_char += 1
            if nom_char != "_":
                index_ocr_word += 1

            if index_ocr_word == length_ocr_line:
                # Write row
                formatted_image = (
                    f"{book_name}_page{boxes[index_ocr_line]['page']:03}.png"
                )
                formatted_id = f"{book_name}.{boxes[index_ocr_line]['page']:03}.{boxes[index_ocr_line]['box_number']:03}"
                worksheet.write(row_index_global, 0, formatted_image, black_text_format)
                worksheet.write(row_index_global, 1, formatted_id, black_text_format)
                worksheet.write(
                    row_index_global,
                    2,
                    json.dumps(boxes[index_ocr_line]["point"]),
                    black_text_format,
                )
                if index_ocr_line >= len(nom_sentences) - 1:
                    # Drain remaining chars
                    while count_char < length_char:
                        nom_char, nom_status = aligned_nom[count_char]
                        qn_char, qn_status = aligned_qn[count_char]
                        nom_pairs.append(
                            red_text_format
                            if nom_status == "red"
                            else black_text_format
                        )
                        nom_pairs.append(nom_char)
                        qn_pairs.append(
                            red_text_format if qn_status == "red" else black_text_format
                        )
                        qn_pairs.append(qn_char + " ")
                        if nom_char == "_" or qn_char == "_":
                            empty_word += 1
                        elif nom_status == "red" and nom_status == "red":
                            red_word += 1
                        total_word += 1
                        count_char += 1
                    write_line(row_index_global, nom_pairs, qn_pairs, length_ocr_line)
                    row_index_global += 1
                    break

                write_line(row_index_global, nom_pairs, qn_pairs, length_ocr_line)
                row_index_global += 1
                index_ocr_line += 1
                index_ocr_word = 0
                nom_pairs = []
                qn_pairs = []
        
    print(f"Total word: {total_word}")
    print(f"Red word: {red_word}")
    print(f"Empty word: {empty_word}")
    print(f"Percentage: {(total_word - red_word - empty_word) / total_word}")
    workbook.close()


# endregion


# region Main function


def processAlignment(
    book_name, input_nom_folder, input_qngu_folder, qn_sino_dict, sino_similar_dict
):
    page_ranges_file = "page_ranges.json"
    output_file = f"output_Sach_012_after_processing.xlsx"

    # Load page ranges from JSON file
    with open(page_ranges_file, "r", encoding="utf-8") as f:
        page_ranges = json.load(f)

    # Check if the book_name exists in the page_ranges
    if book_name not in page_ranges:
        print(f"Book name '{book_name}' not found in page ranges.")
        return

    # Get all page ranges for the specified book
    book_page_ranges = page_ranges[book_name]["page_pairs"]

    write_to_excel(
        input_nom_folder,
        input_qngu_folder,
        book_name,
        book_page_ranges,
        sino_similar_dict,
        qn_sino_dict,
        output_file,
    )


if __name__ == "__main__":
    # Define input and output folders and book name to process
    input_nom_folder = "../OCR_NOM/Sach-Nom-Cong-Giao-1995-012_Processed"
    input_qngu_folder = "../OCR_QN/Sach-Nom-Cong-Giao-1995-012_Postprocessed_Gemini"
    book_name = "Sach-Nom-Cong-Giao-1995-012"

    sino_similar_filename = "SinoNom_similar_Dic.xlsx"
    qn_sino_filename = "Standardized_QuocNgu_SinoNom_Dic.xlsx"

    qn_sino_dict = load_quoc_ngu_sino_nom_dic(qn_sino_filename)
    sino_similar_dict = load_sino_nom_similar_dic(sino_similar_filename)

    processAlignment(
        book_name, input_nom_folder, input_qngu_folder, qn_sino_dict, sino_similar_dict
    )

# endregion
