import json
import os
import re
import numpy as np
import xlsxwriter
import openpyxl
import ast
import json
from preprocess_qn import standardize_vietnamese


def load_sino_nom_similar_dic(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    sino_similar_dict = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
        input_char = row[0]
        similar_chars_str = row[1]
        similar_chars = ast.literal_eval(similar_chars_str) # Convert string to list
        ordered_chars = [input_char] + similar_chars  # Maintain original order
        sino_similar_dict[input_char] = ordered_chars
    return sino_similar_dict


def load_quoc_ngu_sino_nom_dic(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    qn_sino_dict = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        qn_word = row[0]
        sino_char = row[1]

        if qn_word in qn_sino_dict:
            qn_sino_dict[qn_word].append(sino_char)   # Add to list if it exists
        else:
            qn_sino_dict[qn_word] = [sino_char] # Create new list if not
    return qn_sino_dict


def load_json_files_from_folder(folder_path):
    json_data_list = []
    
    filenames = sorted(
        [f for f in os.listdir(folder_path) if f.endswith(".json")],
        key=lambda x: (len(x), x)
    )
    
    print(filenames)

    for filename in filenames:
        file_path = os.path.join(folder_path, filename)
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                json_data = json.load(file)
                json_data_list.append(json_data)
        except Exception as e:
            print(f"Cannot read file {filename}: {e}") 
    
    return json_data_list


def box_alignment(s1, s2):
    aligned_s1 = []
    aligned_s2 = []
    dp = [[0] * (len(s2) + 1) for _ in range(len(s1) + 1)]

    # Fill the DP table for Levenshtein distance
    for i in range(len(s1) + 1):
        for j in range(len(s2) + 1):
            if i == 0:
                dp[i][j] = j  # Cost of insertion
            elif j == 0:
                dp[i][j] = i  # Cost of deletion
            else:
                count_s1 = s1[i - 1]
                count_s2 = s2[j - 1]
                dp[i][j] = min(
                    dp[i - 1][j] + 1,  # Deletion
                    dp[i][j - 1] + 1,  # Insertion
                    dp[i - 1][j - 1] + (0 if count_s1 == count_s2 else 1)  # Substitution (only if counts differ)
                )

    # Backtrack to find aligned sequences
    i, j = len(s1), len(s2)
    while i > 0 or j > 0:
        if i > 0 and dp[i][j] == dp[i - 1][j] + 1:
            aligned_s1.append(s1[i - 1])
            aligned_s2.append(None)
            i -= 1
        elif j > 0 and dp[i][j] == dp[i][j - 1] + 1:
            aligned_s1.append(None)
            aligned_s2.append(s2[j - 1])
            j -= 1
        else:
            aligned_s1.append(s1[i - 1])
            aligned_s2.append(s2[j - 1])
            i -= 1
            j -= 1

    aligned_s1.reverse()
    aligned_s2.reverse()

    return aligned_s1, aligned_s2


def create_ocr_nom_string(paragraph):
    results = []
    for line in paragraph:  # Duyệt từng dòng trong 'ocr_text'
        results.extend(list(line))  # Tách từng ký tự và thêm vào danh sách
    return results

def get_cost(ocr_char, qn_word, sino_similar_dict, qn_sino_dict):
    qn_word = re.sub(r'[^\w\s]', '', qn_word).lower()
    qn_word = standardize_vietnamese(qn_word)

    S1 = sino_similar_dict.get(ocr_char, {ocr_char})
    S2 = qn_sino_dict.get(qn_word, set())
    intersection = [char for char in S1 if char in S2]
    if len(intersection) == 0:
        return 2
    else:
        return 0
    
# def Levenshtein_Align_Line(ocr_sentence, qn_sentence, nom_sino_dict, qn_sino_dict):
#     m, n = len(ocr_sentence), len(qn_sentence)
#     dp = np.zeros((m + 1, n + 1))

#     dp[:, 0] = np.arange(m + 1)
#     dp[0, :] = np.arange(n + 1)

#     for i in range(1, m + 1):
#         for j in range(1, n + 1):
#             cost = get_cost(ocr_sentence[i - 1], qn_sentence[j - 1], nom_sino_dict, qn_sino_dict)
#             dp[i, j] = min(dp[i - 1, j] + 1,
#                            dp[i, j - 1] + 1,
#                            dp[i - 1, j - 1] + cost)

#     aligned_result_ocr = []
#     aligned_result_qn = []

#     i, j = m, n
#     while i > 0 or j > 0:
#         if i > 0 and j > 0:
#             cost = get_cost(ocr_sentence[i - 1], qn_sentence[j - 1], nom_sino_dict, qn_sino_dict)
#             if dp[i, j] == dp[i - 1, j - 1] + cost:
#                 status = "red" if cost >= 1 else "black"
#                 aligned_result_ocr.append((ocr_sentence[i - 1], status))
#                 aligned_result_qn.append((qn_sentence[j - 1], status))
#                 i, j = i - 1, j - 1
#                 continue

#         if i > 0 and (j == 0 or dp[i, j] == dp[i - 1, j] + 1):
#             aligned_result_ocr.append((ocr_sentence[i - 1], "red"))
#             aligned_result_qn.append(('-', "red"))
#             i -= 1
#         elif j > 0:
#             aligned_result_ocr.append(('-', "red"))
#             aligned_result_qn.append((qn_sentence[j - 1], "red"))
#             j -= 1

#     return aligned_result_ocr[::-1], aligned_result_qn[::-1]

def minimum_edit_distance_with_cost(sino_nom_string, quoc_ngu_string, sinonom_similar_dic, quocngu_sinonom_dic):
    m, n = len(sino_nom_string), len(quoc_ngu_string)
    dp = np.zeros((m + 1, n + 1))

    for i in range(1, m + 1):
        dp[i][0] = i
    for j in range(1, n + 1):
        dp[0][j] = j

    aligned_result = []

    for i in range(1, m + 1):
        for j in range(1, n + 1):
            cost = get_cost(
                sino_nom_string[i - 1], quoc_ngu_string[j - 1], sinonom_similar_dic, quocngu_sinonom_dic)
            dp[i][j] = min(dp[i - 1][j] + 1, dp[i][j - 1] +
                           1, dp[i - 1][j - 1] + cost)

    i, j = m, n
    while i > 0 or j > 0:
        if i > 0 and j > 0:
            cost = get_cost(sino_nom_string[i - 1], quoc_ngu_string[j - 1], sinonom_similar_dic, quocngu_sinonom_dic)
            if dp[i][j] == dp[i - 1][j - 1] + cost:
                status = "black" if cost == 0 else "red"
                aligned_result.append((sino_nom_string[i - 1], status))
                i, j = i - 1, j - 1
                continue
        if i > 0 and (j == 0 or dp[i][j] == dp[i - 1][j] + 1):
            aligned_result.append((sino_nom_string[i - 1], "red"))
            i -= 1
        elif j > 0:
            aligned_result.append(('-', "red"))
            j -= 1

    aligned_result.reverse()
    return aligned_result, dp[m][n]


def load_single_json_file(file_path):
    with open(file_path, "r", encoding="utf-8") as f:
        return json.load(f)


def load_single_nom_page(file_path):
    nom_sentences = []
    nom_paragraph = load_single_json_file(file_path)

    if "result" in nom_paragraph:
        for item in nom_paragraph["result"]:
            if "transcription" in item:
                nom_sentences.append(item["transcription"])
    return nom_sentences

def load_single_qngu_page(file_path):
    qn_sentences = []
    qn_paragraph = load_single_json_file(file_path)

    if "ocr_text" in qn_paragraph:
        for line in qn_paragraph["ocr_text"]:
            qn_sentences.extend(line.split())
    return qn_sentences


def extract_qn_sentences(file_path):
    qn_sentences = []

    with open(file_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    ocr_text = data.get("ocr_text", [])
    if ocr_text:
        qn_sentences.extend(ocr_text)

    return qn_sentences


def load_box_nom_page(file_path):
    boxes = []
    data = load_single_json_file(file_path)

    if "result" in data:
        for item in data["result"]:
            if "points" in item:
                boxes.append(item["points"])
    return boxes


def write_to_excel(input_nom_folder, input_qngu_folder, book_name, book_page_ranges, sino_similar_dict, qn_sino_dict, output_file):
    workbook = xlsxwriter.Workbook(output_file)
    worksheet = workbook.add_worksheet()
    
    # Define various formatting styles
    black_text_format   = workbook.add_format({'font_name': 'Nom Na Tong', 'font_size': 12, 'color': 'black'})
    header_format       = workbook.add_format({'font_name': 'Nom Na Tong', 'font_size': 12, 'bold': True, 'align': 'center'})
    green_fill          = workbook.add_format({'font_name': 'Nom Na Tong', 'font_size': 12, 'bg_color': '#00FF00'})
    red_text_format     = workbook.add_format({'font_name': 'Nom Na Tong', 'color': '#FF0000'})
    blue_text_format    = workbook.add_format({'font_name': 'Nom Na Tong', 'color': '#0000FF'})

    # Set column widths
    worksheet.set_column('A:A', 20)
    worksheet.set_column('B:B', 50)
    worksheet.set_column('C:C', 30)
    worksheet.set_column('D:D', 40)

    headers = ["ID", "Image Box", "SinoNom OCR", "Chữ Quốc Ngữ"]
    
    row_index_global = 1
    
    for col, header in enumerate(headers):
        worksheet.write(0, col, header, header_format)
    
    total_word = 0
    red_word = 0
    empty_word = 0

    for page_x, page_y in book_page_ranges:      
        boxes = load_box_nom_page(f"{input_nom_folder}/page_{page_y}.json")
        
        # Lấy câu sino-nom và quốc ngữ tương ứng
        nom_sentences   = load_single_nom_page(f"{input_nom_folder}/page_{page_y}.json")
        qn_sentences    = extract_qn_sentences(f"{input_qngu_folder}/page_{page_x}.json")
        print("Length of qn_sentences: ", len(qn_sentences))
        
        # qn_words_string = load_single_qngu_page(f"{input_qngu_folder}/page_{page_x}.json")        

        nom_lengths = [len(line) for line in nom_sentences]
        qn_lengths = [len(line.split()) for line in qn_sentences]

        aligned_nom, aligned_qn = box_alignment(nom_lengths, qn_lengths)

        print("Nom sentence: ", nom_sentences)
        print("Quoc Ngu: ", qn_sentences)
        
        print("Aligned Nom: ", aligned_nom)
        print("Aligned QN: ", aligned_qn)

        qn_index = 0
        ocr_index = 0
        qn_sentence_index = 0
        
        for row_index in range(len(aligned_qn)):
            print('----'*5)
            print(f"Row index: {row_index}")
            print(f"QN index: {qn_index}")
            print(f"OCR index: {ocr_index}")
            format_pairs = []        
            
            if (qn_index < len(qn_sentences) and ocr_index < len(nom_sentences)):
                formatted_id = f"{book_name}.page{page_x}_page{page_y}.{ocr_index:03}" 
                box_info = json.dumps(boxes[ocr_index])
                matched_qn = qn_sentences[qn_index]
                matched_nom = nom_sentences[ocr_index]
            
                nom_words_string= create_ocr_nom_string(matched_nom)
                qn_words_string = matched_qn.split()

                print("Nom words string: ", nom_words_string)
                print("QN words string: ", qn_words_string)
                
                if aligned_qn[row_index] is None:
                    worksheet.write(row_index, 0, formatted_id, black_text_format)
                    worksheet.write(row_index, 1, box_info, green_fill)
                    worksheet.write(row_index, 2, matched_nom, black_text_format)
                    worksheet.write(row_index, 3, matched_qn, black_text_format)
                    ocr_index += 1
                    continue
                
                if aligned_nom[row_index] is None:
                    worksheet.write(row_index, 0, formatted_id, black_text_format)
                    worksheet.write(row_index, 1, box_info, green_fill)
                    worksheet.write(row_index, 2, matched_nom, black_text_format)
                    worksheet.write(row_index, 3, matched_qn, black_text_format)
                    qn_index += 1
                    continue
                
                worksheet.write(row_index, 0, formatted_id, black_text_format)
                worksheet.write(row_index, 1, box_info, black_text_format)
                worksheet.write(row_index, 2, matched_nom, black_text_format)
                worksheet.write(row_index, 3, matched_qn, black_text_format)
                
                qn_index += 1
                ocr_index += 1
                
                col_index = 2
                
                aligned_result, _ = minimum_edit_distance_with_cost(nom_words_string, qn_words_string, sino_similar_dict, qn_sino_dict)
                print(aligned_result)
                for char, status in aligned_result:
                    if status == "black":
                        format_pairs.append(black_text_format)
                        format_pairs.append(char)
                    elif status == "red":
                        format_pairs.append(red_text_format)
                        format_pairs.append(char)

                qn_sentence_index += 1
                worksheet.write_rich_string(row_index, col_index, *format_pairs)
            elif (ocr_index >= len(nom_sentences) and aligned_nom[row_index] is None):
                matched_nom = "No sentences"
                matched_qn = qn_sentences[qn_index]
                worksheet.write(row_index, 0, "", green_fill)
                worksheet.write(row_index, 1, "", green_fill)
                worksheet.write(row_index, 2, matched_nom, black_text_format)
                worksheet.write(row_index, 3, matched_qn, black_text_format)

                worksheet.write_rich_string(row_index, col_index, *format_pairs)
            else:
                matched_nom = nom_sentences[ocr_index]
                matched_qn = "No sentences"
                formatted_id = f"{book_name}.page{page_x}_page{page_y}.{ocr_index:03}" 
                box_info = json.dumps(boxes[ocr_index])
                worksheet.write(row_index, 0, formatted_id, green_fill)
                worksheet.write(row_index, 1, box_info, green_fill)
                worksheet.write(row_index, 2, matched_nom, black_text_format)
                worksheet.write(row_index, 3, matched_qn, black_text_format)
                
    # print(f"Total word: {total_word}")
    # print(f"Red word: {red_word}")
    # print(f"Empty word: {empty_word}")
    # print(f"Percentage: {(total_word - red_word - empty_word)/total_word}")
    
    workbook.close()


def processAlignment(book_name, qn_sino_dict, sino_similar_dict):
    input_nom_folder    = r"D:\CS418-Final-Project\src\Output_OCR_Nom_Sach_001_Processed"
    input_qngu_folder   = r"D:\CS418-Final-Project\src\Output_OCR_QN_Sach_001_Processed"
    page_ranges_file    = "../extracted_imgs/page_ranges.json"
    output_file         = "output.xlsx"
    output_file         = f"output_{book_name}.xlsx"

    # Load page ranges from JSON file
    with open(page_ranges_file, "r", encoding="utf-8") as f:
        page_ranges = json.load(f)

    # Check if the book_name exists in the page_ranges
    if book_name not in page_ranges:
        print(f"Book name '{book_name}' not found in page ranges.")
        return

    # Get all page ranges for the specified book
    book_page_ranges = page_ranges[book_name]["page_pairs"]
    
    write_to_excel(input_nom_folder, input_qngu_folder, book_name, book_page_ranges, sino_similar_dict, qn_sino_dict, output_file)


if __name__ == "__main__":
    sino_similar_filename = "SinoNom_similar_Dic.xlsx"
    qn_sino_filename = "Standardized_QuocNgu_SinoNom_Dic.xlsx"

    qn_sino_dict = load_quoc_ngu_sino_nom_dic(qn_sino_filename)
    sino_similar_dict = load_sino_nom_similar_dic(sino_similar_filename)
    # print( get_cost('秩', 'ni', sino_similar_dict, qn_sino_dict))
    processAlignment('Sach-Nom-Cong-Giao-1995-001', qn_sino_dict, sino_similar_dict)