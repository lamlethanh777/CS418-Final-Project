import json
import os
import re
import xlsxwriter
import openpyxl
import ast
import json

def load_sino_nom_similar_dic(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    sino_similar_dict = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
        input_char = row[0]
        similar_chars_str = row[1]
        similar_chars = ast.literal_eval(similar_chars_str) # Convert string to list
        ordered_chars = [input_char] + similar_chars  # Maintain original order
        sino_similar_dict[input_char] = ordered_chars
    return sino_similar_dict
        
def load_quoc_ngu_sino_nom_dic(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    qn_sino_dict = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        qn_word = row[0]
        sino_char = row[1]

        if qn_word in qn_sino_dict:
            qn_sino_dict[qn_word].append(sino_char)   # Add to list if it exists
        else:
            qn_sino_dict[qn_word] = [sino_char] # Create new list if not
    return qn_sino_dict

def load_json_files_from_folder(folder_path):
    json_data_list = []
    # Lấy danh sách tệp và sắp xếp theo tên
    filenames = sorted(
        [f for f in os.listdir(folder_path) if f.endswith(".json")]
    )
    for filename in filenames:
        file_path = os.path.join(folder_path, filename)
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                json_data = json.load(file)
                json_data_list.append(json_data)
        except Exception as e:
            print(f"Cannot read file {filename}: {e}") 
    return json_data_list

def write_to_excel(nom_sentences ,qn_words_string, sino_similar_dict, qn_sino_dict, output_file):
    workbook = xlsxwriter.Workbook(output_file)
    worksheet = workbook.add_worksheet()
    
    # Define various formatting styles
    default_font = workbook.add_format({'font_name': 'Nom Na Tong', 'font_size': 12, 'color': 'black'})
    header_format = workbook.add_format({'font_name': 'Nom Na Tong', 'font_size': 12, 'bold': True, 'align': 'center'})
    green_fill = workbook.add_format({'font_name': 'Nom Na Tong', 'font_size': 12, 'bg_color': '#00FF00'})
    red_text_format = workbook.add_format({'font_name': 'Nom Na Tong', 'color': '#FF0000'})
    blue_text_format = workbook.add_format({'font_name': 'Nom Na Tong', 'color': '#0000FF'})

    # Set column widths
    worksheet.set_column('A:A', 20)
    worksheet.set_column('B:B', 50)
    worksheet.set_column('C:C', 30)
    worksheet.set_column('D:D', 40)

    headers = ["ID", "Image Box", "SinoNom OCR", "Chữ Quốc Ngữ"]
    for col, header in enumerate(headers):
        worksheet.write(0, col, header, header_format)

    for i, nom_sentence in enumerate(nom_sentences):
        if "ocr_text" in nom_sentence:
            for j, line in enumerate(nom_sentence["ocr_text"]):
                print(line)
    


def processAlignment(qn_sino_dict, sino_similar_dict):
    input_nom_folder = "Output_OCR_Nom"
    output_qngu_folder = "Output_OCR_QN"
    output_file = "output.xlsx"
    nom_sentences = load_json_files_from_folder(input_nom_folder)
    qngu_sentences = load_json_files_from_folder(output_qngu_folder)
    qn_words_string = []
    for qn_sentence in qngu_sentences:
        if "ocr_text" in qn_sentence:  # Kiểm tra trường ocr_text có tồn tại
            for line in qn_sentence["ocr_text"]:  # Duyệt từng dòng trong ocr_text
                qn_words_string.extend(line.split())  # Tách dòng thành các từ và thêm vào danh sách
    print(qn_words_string)


if __name__ == "__main__":
    sino_similar_filename = "SinoNom_similar_Dic.xlsx"
    qn_sino_filename = "QuocNgu_SinoNom_Dic.xlsx"

    qn_sino_dict = load_quoc_ngu_sino_nom_dic(qn_sino_filename)
    sino_similar_dict = load_sino_nom_similar_dic(sino_similar_filename)
    processAlignment(qn_sino_dict, sino_similar_dict)