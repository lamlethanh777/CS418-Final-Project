import json
import os
import re
import numpy as np
import xlsxwriter
import openpyxl
import ast
import json

def load_sino_nom_similar_dic(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    sino_similar_dict = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
        input_char = row[0]
        similar_chars_str = row[1]
        similar_chars = ast.literal_eval(similar_chars_str) # Convert string to list
        ordered_chars = [input_char] + similar_chars  # Maintain original order
        sino_similar_dict[input_char] = ordered_chars
    return sino_similar_dict
        
def load_quoc_ngu_sino_nom_dic(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    qn_sino_dict = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        qn_word = row[0]
        sino_char = row[1]

        if qn_word in qn_sino_dict:
            qn_sino_dict[qn_word].append(sino_char)   # Add to list if it exists
        else:
            qn_sino_dict[qn_word] = [sino_char] # Create new list if not
    return qn_sino_dict

def load_json_files_from_folder(folder_path):
    json_data_list = []
    # Lấy danh sách tệp và sắp xếp theo tên
    filenames = sorted(
        [f for f in os.listdir(folder_path) if f.endswith(".json")]
    )
    for filename in filenames:
        file_path = os.path.join(folder_path, filename)
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                json_data = json.load(file)
                json_data_list.append(json_data)
        except Exception as e:
            print(f"Cannot read file {filename}: {e}") 
    return json_data_list

def create_ocr_nom_string(paragraph):
    results = []
    for line in paragraph:  # Duyệt từng dòng trong 'ocr_text'
        results.extend(list(line))  # Tách từng ký tự và thêm vào danh sách
    return results

def get_cost(ocr_char, qn_word, sino_similar_dict, qn_sino_dict):
    qn_word = re.sub(r'[^\w\s]', '', qn_word).lower()
    print(qn_word)
    S1 = sino_similar_dict.get(ocr_char, {ocr_char})
    S2 = qn_sino_dict.get(qn_word, set())
    intersection = [char for char in S1 if char in S2]
    if len(intersection) == 0:
        return 1
    else:
        return 0
    
def Levenshtein_Align_Line(ocr_sentence, qn_sentence, nom_sino_dict, qn_sino_dict):
    m, n = len(ocr_sentence), len(qn_sentence)
    dp = np.zeros((m + 1, n + 1))

    dp[:, 0] = np.arange(m + 1)
    dp[0, :] = np.arange(n + 1)

    # Tính toán chi phí chỉnh sửa
    for i in range(1, m + 1):
        for j in range(1, n + 1):
            cost = get_cost(ocr_sentence[i - 1], qn_sentence[j - 1], nom_sino_dict, qn_sino_dict)
            dp[i, j] = min(dp[i - 1, j] + 1,
                           dp[i, j - 1] + 1,
                           dp[i - 1, j - 1] + cost)

    aligned_result_ocr = []
    aligned_result_qn = []

    i, j = m, n
    while i > 0 or j > 0:
        if i > 0 and j > 0:
            cost = get_cost(ocr_sentence[i - 1], qn_sentence[j - 1], nom_sino_dict, qn_sino_dict)
            if dp[i, j] == dp[i - 1, j - 1] + cost:
                status = "red" if cost == 1 else "black"
                aligned_result_ocr.append((ocr_sentence[i - 1], status))
                aligned_result_qn.append((qn_sentence[j - 1], status))
                i, j = i - 1, j - 1
                continue

        if i > 0 and (j == 0 or dp[i, j] == dp[i - 1, j] + 1):
            aligned_result_ocr.append((ocr_sentence[i - 1], "red"))
            aligned_result_qn.append(('_', "red"))
            i -= 1
        elif j > 0:  # Chèn trong OCR
            aligned_result_ocr.append(('_', "red"))
            aligned_result_qn.append((qn_sentence[j - 1], "red"))
            j -= 1

    return aligned_result_ocr[::-1], aligned_result_qn[::-1]  # Đảo ngược để đúng thứ tự


def write_to_excel(nom_sentences ,qn_words_string, sino_similar_dict, qn_sino_dict, output_file):
    workbook = xlsxwriter.Workbook(output_file)
    worksheet = workbook.add_worksheet()
    
    # Define various formatting styles
    black_text_format = workbook.add_format({'font_name': 'Nom Na Tong', 'font_size': 12, 'color': 'black'})
    header_format = workbook.add_format({'font_name': 'Nom Na Tong', 'font_size': 12, 'bold': True, 'align': 'center'})
    green_fill = workbook.add_format({'font_name': 'Nom Na Tong', 'font_size': 12, 'bg_color': '#00FF00'})
    red_text_format = workbook.add_format({'font_name': 'Nom Na Tong', 'color': '#FF0000'})
    blue_text_format = workbook.add_format({'font_name': 'Nom Na Tong', 'color': '#0000FF'})

    # Set column widths
    worksheet.set_column('A:A', 20)
    worksheet.set_column('B:B', 50)
    worksheet.set_column('C:C', 30)
    worksheet.set_column('D:D', 40)

    headers = ["ID", "Image Box", "SinoNom OCR", "Chữ Quốc Ngữ"]
    row_index_global = 1
    
    for col, header in enumerate(headers):
        worksheet.write(0, col, header, header_format)

    nom_words_string = []

    # for paragraph in reversed(nom_paragraphs):
    #     if "ocr_text" in paragraph:
    #         nom_words_string= create_ocr_nom_string(paragraph)
    #         aligned_ocrLinesString, aligned_qnLinesString = Levenshtein_Align_Line(nom_words_string, qn_words_string)
    #         print(aligned_ocrLinesString)
    #         print(aligned_qnLinesString)

    nom_words_string= create_ocr_nom_string(nom_sentences)

    aligned_nomLinesString, aligned_qnLinesString = Levenshtein_Align_Line(nom_words_string, qn_words_string, sino_similar_dict, qn_sino_dict)
        
    # Xử lý ghi nội dung vào hai cột SinoNom OCR và Chữ Quốc Ngữ
    nom_format_pairs = []
    qn_format_pairs = []

    index_ocr_line = 0
    index_ocr_word = 0

    length_ocr_line = len(nom_sentences[index_ocr_line])


    worksheet.write(row_index_global, 0, '1', black_text_format)
    worksheet.write(row_index_global, 1, 'tiendayy', black_text_format)

    for (nom_char, nom_status), (qn_char, qn_status) in zip(aligned_nomLinesString, aligned_qnLinesString):
        # Xử lý định dạng màu cho SinoNom
        if (nom_char != '_'): index_ocr_word += 1
        if nom_status == "red":
            nom_format_pairs.append(red_text_format)
            nom_format_pairs.append(nom_char)
            
        else:
            nom_format_pairs.append(black_text_format)
            nom_format_pairs.append(nom_char)
    
        # Xử lý định dạng màu cho Chữ Quốc Ngữ
        if qn_status == "red":
            qn_format_pairs.append(red_text_format)
            qn_format_pairs.append(qn_char + ' ')
        else:
            qn_format_pairs.append(black_text_format)
            qn_format_pairs.append(qn_char + ' ')
        
        if (index_ocr_word == length_ocr_line):
            qn_format_pairs[-1] = qn_format_pairs[-1].strip()
            if (length_ocr_line == 1):
                worksheet.write(row_index_global, 2, nom_format_pairs[1], nom_format_pairs[0])
                worksheet.write(row_index_global, 3, qn_format_pairs[1], qn_format_pairs[0])
            else:
                worksheet.write_rich_string(row_index_global, 2, *nom_format_pairs)
                worksheet.write_rich_string(row_index_global, 3, *qn_format_pairs)
            index_ocr_line += 1
            index_ocr_word = 0

            row_index_global += 1
            if (index_ocr_line) >= len(nom_sentences): break
            length_ocr_line = len(nom_sentences[index_ocr_line])

            nom_format_pairs = []
            qn_format_pairs = []
    workbook.close()



def processAlignment(qn_sino_dict, sino_similar_dict):
    input_nom_folder = "Output_OCR_Nom"
    output_qngu_folder = "Output_OCR_QN"
    output_file = "output.xlsx"
    nom_paragraph = load_json_files_from_folder(input_nom_folder)
    nom_sentences = []
    for paragraph in reversed(nom_paragraph):
        if "ocr_text" in paragraph:
            nom_sentences.extend(paragraph["ocr_text"])

    qngu_sentences = load_json_files_from_folder(output_qngu_folder)
    qn_words_string = []
    for qn_sentence in qngu_sentences:
        if "ocr_text" in qn_sentence:  # Kiểm tra trường ocr_text có tồn tại
            for line in qn_sentence["ocr_text"]:  # Duyệt từng dòng trong ocr_text
                qn_words_string.extend(line.split())  # Tách dòng thành các từ và thêm vào danh sách
    
    write_to_excel(nom_sentences, qn_words_string, sino_similar_dict, qn_sino_dict, output_file)


if __name__ == "__main__":
    sino_similar_filename = "SinoNom_similar_Dic.xlsx"
    qn_sino_filename = "QuocNgu_SinoNom_Dic.xlsx"

    qn_sino_dict = load_quoc_ngu_sino_nom_dic(qn_sino_filename)
    sino_similar_dict = load_sino_nom_similar_dic(sino_similar_filename)
    processAlignment(qn_sino_dict, sino_similar_dict)