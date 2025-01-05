import json
import os
import re
import numpy as np
import xlsxwriter
import openpyxl
import ast
import json

# BASIC FUNCTIONS
# Sino - QN Punctuations
# Sino - QN Characters

nom_punctuation = set('、。.？！：；')
qn_punctuation = set(',.?!:;')
similar_punctuation = {
    '，': ',', '。': '.', '？': '?', '！': '!', '：': ':', '；': ';'
}

number = set('0123456789')

sino_nom_ranges = [
        (0x3400, 0x4DBF),  # CJK Extension A
        (0x4E00, 0x9FFF),  # CJK Unified Ideographs
        (0xF900, 0xFAFF),  # CJK Compatibility Ideographs
        (0x20000, 0x2A6DF),  # CJK Extension B
        (0x2A700, 0x2B73F),  # CJK Extension C
        (0x2B740, 0x2B81F),  # CJK Extension D
        (0x2B820, 0x2CEAF),  # CJK Extension E
        (0x2CEB0, 0x2EBEF),  # CJK Extension F
        (0x30000, 0x3134F),  # CJK Extension G
    ]

vietnamese_char_pattern = re.compile(
    r"[A-Za-zÀ-ÃÈ-ÊÌ-ÍÒ-ÕÙ-ÚÝà-ãè-êì-íò-õù-úýĂăĐđĨĩŨũƠơƯư\u1EA0-\u1EF9]"
)

def is_nom_char(ch):
    return any(start <= ord(ch) <= end for start, end in sino_nom_ranges)

def is_vietnamese_char(ch):
    return vietnamese_char_pattern.match(ch)


def load_sino_nom_similar_dic(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    sino_similar_dict = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
        input_char = row[0]
        similar_chars_str = row[1]
        similar_chars = ast.literal_eval(similar_chars_str) # Convert string to list
        ordered_chars = [input_char] + similar_chars  # Maintain original order
        sino_similar_dict[input_char] = ordered_chars
    return sino_similar_dict
        
def load_quoc_ngu_sino_nom_dic(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    qn_sino_dict = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        qn_word = row[0]
        sino_char = row[1]

        if qn_word in qn_sino_dict:
            qn_sino_dict[qn_word].append(sino_char)   # Add to list if it exists
        else:
            qn_sino_dict[qn_word] = [sino_char] # Create new list if not

    return qn_sino_dict


def load_parallel_files_from_folder(folder_path):
    parallel_files = []
    json_files = [filename for filename in os.listdir(folder_path) if filename.endswith('.json')]
    
    json_files.sort(key=lambda f: int(f.split('_')[1].split('.')[0]))
    
    for filename in json_files:
        with open(os.path.join(folder_path, filename), 'r', encoding='utf-8') as file:
            parallel_files.append(json.load(file))
    
    return parallel_files

def create_ocr_nom_words_list(paragraph):
    results = []
    all_chars = list(paragraph)
    l = len(all_chars)
    i = 0
    while i < l:
        if all_chars[i] == ' ':
            i += 1
            continue
        if all_chars[i] in number:
            tmp = all_chars[i]
            j = i + 1
            while j < l and (all_chars[j] in number or all_chars[j] in nom_punctuation):
                tmp += all_chars[j]
                j += 1
            results.append(tmp)
            i = j
        elif is_vietnamese_char(all_chars[i]):
            tmp = all_chars[i]
            j = i + 1
            while j < l and (is_vietnamese_char(all_chars[j]) or all_chars[j] in nom_punctuation):
                tmp += all_chars[j]
                j += 1
            results.append(tmp)
            i = j
        elif i < l - 1 and all_chars[i + 1] in nom_punctuation:
            results.append(all_chars[i] + all_chars[i + 1])
            i += 2
        else:
            results.append(all_chars[i])
            i += 1
            
    return results

def create_qn_words_list(qn_string):
    return qn_string.split()

def get_cost(ocr_char, qn_word, sino_similar_dict, qn_sino_dict, standardizer):
    # Case of similar numbers
    if ocr_char == qn_word: 
        return 0
    
    # Case of similar sino characters
    qn_word = ''.join([ch for ch in qn_word if is_vietnamese_char(ch)]).lower()
    ocr_char = ''.join([ch for ch in ocr_char if is_nom_char(ch)])
    if ocr_char == qn_word:
        return 0
    
    standardized_qn_word = standardizer.standardize_vietnamese(qn_word)
    # print(ocr_char, "-", qn_word, "-", standardized_qn_word)

    S1 = sino_similar_dict.get(ocr_char, {ocr_char})
    S2 = qn_sino_dict.get(standardized_qn_word, set())
    
    intersection = [char for char in S1 if char in S2]
    if intersection:
        return 0
    return 1
    
def Levenshtein_Align_Line(ocr_sentence, qn_sentence, nom_sino_dict, qn_sino_dict, standardizer):
    m, n = len(ocr_sentence), len(qn_sentence)
    dp = np.zeros((m + 1, n + 1))

    dp[:, 0] = np.arange(m + 1)
    dp[0, :] = np.arange(n + 1)

    # Tính toán chi phí chỉnh sửa
    for i in range(1, m + 1):
        for j in range(1, n + 1):
            cost = get_cost(ocr_sentence[i - 1], qn_sentence[j - 1], nom_sino_dict, qn_sino_dict, standardizer)
            dp[i, j] = min(dp[i - 1, j] + 1,
                           dp[i, j - 1] + 1,
                           dp[i - 1, j - 1] + cost)

    aligned_result_ocr = []
    aligned_result_qn = []

    i, j = m, n
    while i > 0 or j > 0:
        if i > 0 and j > 0:
            cost = get_cost(ocr_sentence[i - 1], qn_sentence[j - 1], nom_sino_dict, qn_sino_dict, standardizer)
            if dp[i, j] == dp[i - 1, j - 1] + cost:
                status = "red" if cost == 1 else "black"
                aligned_result_ocr.append((ocr_sentence[i - 1], status))
                aligned_result_qn.append((qn_sentence[j - 1], status))
                i, j = i - 1, j - 1
                continue

        if i > 0 and (j == 0 or dp[i, j] == dp[i - 1, j] + 1):
            aligned_result_ocr.append((ocr_sentence[i - 1], "red"))
            aligned_result_qn.append(('_', "red"))
            i -= 1
        elif j > 0:  # Chèn trong OCR
            aligned_result_ocr.append(('_', "red"))
            aligned_result_qn.append((qn_sentence[j - 1], "red"))
            j -= 1

    return aligned_result_ocr[::-1], aligned_result_qn[::-1]  # Đảo ngược để đúng thứ tự


def write_one_pair_to_excel(nom_sentences, qn_words_string, sino_similar_dict, qn_sino_dict, standardizer, worksheet, black_text_format, red_text_format, row_index_global):
    nom_words_list= create_ocr_nom_words_list(nom_sentences)
    qn_words_list = create_qn_words_list(qn_words_string)

    aligned_nomLinesString, aligned_qnLinesString = Levenshtein_Align_Line(nom_words_list, qn_words_list, sino_similar_dict, qn_sino_dict, standardizer)
        
    # Xử lý ghi nội dung vào hai cột SinoNom OCR và Chữ Quốc Ngữ
    nom_format_pairs = []
    qn_format_pairs = []

    worksheet.write(row_index_global, 0, '', black_text_format)
    worksheet.write(row_index_global, 1, '', black_text_format)

    for (nom_char, nom_status), (qn_char, qn_status) in zip(aligned_nomLinesString, aligned_qnLinesString):
        if nom_status == "red":
            nom_format_pairs.append(red_text_format)
            nom_format_pairs.append("<" + nom_char + ">")
            
        else:
            nom_format_pairs.append(black_text_format)
            nom_format_pairs.append(nom_char)
    
        if qn_status == "red":
            qn_format_pairs.append(red_text_format)
            qn_format_pairs.append("<" + qn_char + "> ")
        else:
            qn_format_pairs.append(black_text_format)
            qn_format_pairs.append(qn_char + ' ')
            
    if len(nom_format_pairs) > 2:
        worksheet.write_rich_string(row_index_global, 0, *nom_format_pairs)
        worksheet.write_rich_string(row_index_global, 1, *qn_format_pairs)
    else:
        worksheet.write(row_index_global, 0, nom_format_pairs[1], nom_format_pairs[0])
        worksheet.write(row_index_global, 1, qn_format_pairs[1], qn_format_pairs[0])

def write_to_excel(page, sino_similar_dict, qn_sino_dict, standardizer, output_file):
    workbook = xlsxwriter.Workbook(output_file)
    worksheet = workbook.add_worksheet()
    
    # Define various formatting styles
    black_text_format = workbook.add_format({'font_name': 'Nom Na Tong', 'font_size': 12, 'color': 'black'})
    header_format = workbook.add_format({'font_name': 'Nom Na Tong', 'font_size': 12, 'bold': True, 'align': 'center'})
    green_fill = workbook.add_format({'font_name': 'Nom Na Tong', 'font_size': 12, 'bg_color': '#00FF00'})
    red_text_format = workbook.add_format({'font_name': 'Nom Na Tong', 'color': '#FF0000'})
    blue_text_format = workbook.add_format({'font_name': 'Nom Na Tong', 'color': '#0000FF'})

    # Set column widths
    worksheet.set_column('A:A', 50)
    worksheet.set_column('B:B', 100)


    headers = ["SinoNom", "Chữ Quốc Ngữ"]
    
    for col, header in enumerate(headers):
        worksheet.write(0, col, header, header_format)

    row_index_global = 1
    if len(page["title"]) == 2:
        write_one_pair_to_excel(page["title"][0], page["title"][1], sino_similar_dict, qn_sino_dict, standardizer, worksheet, black_text_format, red_text_format, row_index_global)
        row_index_global += 1

    for i, (nom_sentences, qn_words_string) in enumerate(page["content"], row_index_global):
        write_one_pair_to_excel(nom_sentences, qn_words_string, sino_similar_dict, qn_sino_dict, standardizer, worksheet, black_text_format, red_text_format, i)
                
    workbook.close()
    print(f"File {output_file} has been written successfully.")


def processAlignment(qn_sino_dict, sino_similar_dict, standardizer):
    parallel_files = load_parallel_files_from_folder("test")
    
    if not os.path.exists("output"):
        os.makedirs("output")
    for i, file in enumerate(parallel_files):
        output_file = "output/" + str(i + 1) + ".xlsx"
        write_to_excel(file, sino_similar_dict, qn_sino_dict, standardizer, output_file)


from qn_standardizer import QNStandardizer

if __name__ == "__main__":
    sino_similar_filename = "SinoNom_similar_Dic.xlsx"
    qn_sino_filename = "Standardized_QuocNgu_SinoNom_Dic.xlsx"
    
    standardizer = QNStandardizer()

    qn_sino_dict = load_quoc_ngu_sino_nom_dic(qn_sino_filename)
    sino_similar_dict = load_sino_nom_similar_dic(sino_similar_filename)
    processAlignment(qn_sino_dict, sino_similar_dict, standardizer)